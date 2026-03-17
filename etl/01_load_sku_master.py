from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from common import (
    connect,
    finish_etl_run,
    get_config,
    print_banner,
    record_file_import,
    register_etl_run,
    utc_now_iso,
)


SOURCE_FILE = "99_SKU_MASTER.xlsx"


def normalize_text(value: object) -> str:
    return str(value).strip() if value is not None else ""


def load_rows(source_path: Path) -> list[tuple[str, str]]:
    wb = load_workbook(source_path, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    rows: list[tuple[str, str]] = []
    for sku, product_name, *_ in ws.iter_rows(min_row=2, values_only=True):
        sku_text = normalize_text(sku)
        product_name_text = normalize_text(product_name)
        if not sku_text:
            continue
        if not product_name_text:
            raise ValueError(f"SKU {sku_text} is missing product name in {source_path.name}")
        rows.append((sku_text, product_name_text))
    return rows


def main() -> int:
    config = get_config()
    source_path = config.base_dir / SOURCE_FILE
    if not source_path.exists():
        raise FileNotFoundError(f"Source file not found: {source_path}")

    print_banner(f"Loading SKU master from {source_path.name}")
    conn = connect(config.db_path)
    run_id = register_etl_run(
        conn,
        script_name="01_load_sku_master.py",
        run_type="load_sku_master",
        status="started",
    )

    try:
        rows = load_rows(source_path)
        record_file_import(
            conn,
            run_id=run_id,
            source_file=source_path,
            file_role="sku_master",
            import_status="loaded",
            row_count=len(rows),
        )

        conn.executemany(
            """
            INSERT INTO dim_sku (
                sku,
                product_name_cn,
                created_at,
                updated_at
            ) VALUES (?, ?, ?, ?)
            ON CONFLICT(sku) DO UPDATE SET
                product_name_cn = excluded.product_name_cn,
                updated_at = excluded.updated_at
            """,
            [(sku, product_name, utc_now_iso(), utc_now_iso()) for sku, product_name in rows],
        )

        alias_rows = []
        for sku, product_name in rows:
            alias_value = product_name.strip().lower()
            alias_rows.append(("product_name_cn", alias_value, sku, 1, utc_now_iso()))

        conn.executemany(
            """
            INSERT INTO dim_sku_alias (
                alias_type,
                alias_value,
                sku,
                is_unique_mapping,
                created_at
            ) VALUES (?, ?, ?, ?, ?)
            ON CONFLICT(alias_type, alias_value, sku) DO UPDATE SET
                is_unique_mapping = excluded.is_unique_mapping
            """,
            alias_rows,
        )

        conn.commit()
        finish_etl_run(conn, run_id, "success", f"Loaded {len(rows)} SKU master rows.")
        conn.commit()
        print_banner(f"Loaded {len(rows)} SKU master rows")
        return 0
    except Exception as exc:  # pragma: no cover
        finish_etl_run(conn, run_id, "failed", str(exc))
        conn.commit()
        raise
    finally:
        conn.close()


if __name__ == "__main__":
    raise SystemExit(main())
