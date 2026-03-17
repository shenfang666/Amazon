from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from common import (
    connect,
    finish_etl_run,
    get_config,
    print_banner,
    record_file_import,
    register_etl_run,
    utc_now_iso,
)


SOURCE_FILE = "98_SKU Cost Table_Amazon.xlsx"
SOURCE_SHEET = "2.9 SKU Cost Table"


def normalize_text(value: object) -> str:
    return str(value).strip() if value is not None else ""


def normalize_month(value: object) -> str:
    text = normalize_text(value)
    if not text:
        return ""
    return text[:7]


def to_number(value: object) -> float:
    if value is None or value == "":
        raise ValueError("Numeric field is empty")
    if isinstance(value, (int, float)):
        return float(value)
    return float(str(value).strip())


def load_cost_rows(source_path: Path) -> list[tuple[str, str, float, float, str]]:
    wb = load_workbook(source_path, data_only=True, read_only=True)
    ws = wb[SOURCE_SHEET]
    rows: list[tuple[str, str, float, float, str]] = []
    for idx, row in enumerate(ws.iter_rows(min_row=3, values_only=True), start=3):
        month_raw, product_name_raw, product_cost_raw, inbound_cost_raw, *_ = row
        cost_month = normalize_month(month_raw)
        product_name = normalize_text(product_name_raw)
        if not cost_month and not product_name:
            continue
        if not cost_month or not product_name:
            continue
        product_cost = to_number(product_cost_raw)
        inbound_cost = to_number(inbound_cost_raw)
        rows.append((cost_month, product_name, product_cost, inbound_cost, f"{SOURCE_SHEET}:{idx}"))
    return rows


def main() -> int:
    config = get_config()
    source_path = config.base_dir / SOURCE_FILE
    if not source_path.exists():
        raise FileNotFoundError(f"Source file not found: {source_path}")

    print_banner(f"Loading SKU cost from {source_path.name}")
    conn = connect(config.db_path)
    run_id = register_etl_run(
        conn,
        script_name="02_load_sku_cost.py",
        run_type="load_sku_cost",
        status="started",
    )

    try:
        cost_rows = load_cost_rows(source_path)
        record_file_import(
            conn,
            run_id=run_id,
            source_file=source_path,
            file_role="sku_cost",
            import_status="loaded",
            row_count=len(cost_rows),
        )

        conn.execute(
            "delete from pending_mapping_queue where source_table = 'dim_cost_monthly' and source_file = ?",
            (str(source_path),),
        )

        alias_map = {
            row[0]: row[1]
            for row in conn.execute(
                """
                SELECT alias_value, sku
                FROM dim_sku_alias
                WHERE alias_type = 'product_name_cn'
                  AND is_unique_mapping = 1
                """
            ).fetchall()
        }
        manual_alias_map = {
            row[0]: row[1]
            for row in conn.execute(
                """
                SELECT alias_value, sku
                FROM manual_sku_alias
                WHERE alias_type = 'product_name_cn'
                  AND is_active = 1
                """
            ).fetchall()
        }

        insert_rows: list[tuple[str, str, float, float, str, str, str]] = []
        pending_rows: list[tuple[str, str, str | None, str, str, str, str, str, str | None]] = []

        for cost_month, product_name, product_cost, inbound_cost, row_ref in cost_rows:
            alias_value = product_name.lower()
            sku = manual_alias_map.get(alias_value) or alias_map.get(alias_value)
            if not sku:
                pending_rows.append(
                    (
                        "dim_cost_monthly",
                        str(source_path),
                        None,
                        product_name,
                        "product_name_cn",
                        "pending",
                        f"Cost row {row_ref} could not be mapped to SKU",
                        utc_now_iso(),
                        None,
                    )
                )
                continue

            insert_rows.append(
                (
                    sku,
                    cost_month,
                    product_cost,
                    inbound_cost,
                    str(source_path),
                    row_ref,
                    utc_now_iso(),
                )
            )

        conn.executemany(
            """
            INSERT INTO dim_cost_monthly (
                sku,
                cost_month,
                product_unit_cost,
                inbound_unit_cost,
                source_file,
                source_row_ref,
                created_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?)
            ON CONFLICT(sku, cost_month) DO UPDATE SET
                product_unit_cost = excluded.product_unit_cost,
                inbound_unit_cost = excluded.inbound_unit_cost,
                source_file = excluded.source_file,
                source_row_ref = excluded.source_row_ref
            """,
            insert_rows,
        )

        if pending_rows:
            conn.executemany(
                """
                INSERT INTO pending_mapping_queue (
                    source_table,
                    source_file,
                    source_row_hash,
                    ambiguous_value,
                    mapping_type,
                    status,
                    notes,
                    created_at,
                    resolved_at
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                pending_rows,
            )

        conn.commit()
        note = f"Loaded {len(insert_rows)} cost rows; pending mappings={len(pending_rows)}"
        finish_etl_run(conn, run_id, "success", note)
        conn.commit()
        print_banner(note)
        return 0
    except Exception as exc:  # pragma: no cover
        finish_etl_run(conn, run_id, "failed", str(exc))
        conn.commit()
        raise
    finally:
        conn.close()


if __name__ == "__main__":
    raise SystemExit(main())
