from __future__ import annotations

from datetime import datetime, timedelta
from pathlib import Path

from openpyxl import load_workbook

from common import (
    connect,
    finish_etl_run,
    get_config,
    print_banner,
    record_file_import,
    register_etl_run,
    sha256_text,
    stable_json,
    utc_now_iso,
)


SOURCE_FILE = '5_Test Orders_Amazon.xlsx'
SOURCE_SHEET = 'Sheet1'


def normalize_text(value: object) -> str:
    return str(value).strip() if value is not None else ''


def excel_serial_to_date(value: float) -> str:
    base = datetime(1899, 12, 30)
    dt = base + timedelta(days=float(value))
    return dt.date().isoformat()


def parse_order_date(value: object) -> str | None:
    if value is None or value == '':
        return None
    if isinstance(value, (int, float)):
        return excel_serial_to_date(float(value))
    if isinstance(value, datetime):
        return value.date().isoformat()
    text = normalize_text(value)
    return text or None


def parse_float(value: object) -> float | None:
    text = normalize_text(value)
    if text == '':
        return None
    return float(text)


def fuzzy_product_name_skus(conn, product_name_text: str) -> list[str]:
    keyword = product_name_text.strip().lower()
    if not keyword:
        return []
    rows = conn.execute(
        """
        SELECT DISTINCT sku
        FROM dim_sku
        WHERE instr(lower(product_name_cn), ?) > 0
           OR instr(?, lower(product_name_cn)) > 0
        ORDER BY sku
        """,
        (keyword, keyword),
    ).fetchall()
    return [row[0] for row in rows if row[0]]


def source_order_skus(conn, order_id_text: str) -> list[str]:
    settlement_rows = conn.execute(
        """
        SELECT DISTINCT sku
        FROM fact_settlement_lines
        WHERE order_id = ?
          AND sku IS NOT NULL
        """,
        (order_id_text,),
    ).fetchall()
    order_rows = conn.execute(
        """
        SELECT DISTINCT sku
        FROM fact_order_lines
        WHERE amazon_order_id = ?
          AND sku IS NOT NULL
        """,
        (order_id_text,),
    ).fetchall()
    values = sorted({row[0] for row in settlement_rows + order_rows if row[0]})
    return values


def main() -> int:
    config = get_config()
    source_path = config.base_dir / SOURCE_FILE
    if not source_path.exists():
        raise FileNotFoundError(f'Source file not found: {source_path}')

    print_banner(f'Loading review orders from {source_path.name}')
    conn = connect(config.db_path)
    run_id = register_etl_run(
        conn,
        script_name='06_load_review_orders.py',
        run_type='load_review_orders',
        status='started',
    )

    try:
        wb = load_workbook(source_path, data_only=True, read_only=True)
        ws = wb[SOURCE_SHEET]
        raw_rows = list(ws.iter_rows(min_row=2, values_only=True))

        record_file_import(
            conn,
            run_id=run_id,
            source_file=source_path,
            file_role='review_orders',
            import_status='loaded',
            row_count=len(raw_rows),
        )

        conn.execute(
            "delete from pending_mapping_queue where source_table = 'fact_review_orders' and source_file = ?",
            (str(source_path),),
        )
        conn.execute('DELETE FROM fact_review_orders WHERE source_file = ?', (str(source_path),))

        alias_map: dict[str, list[str]] = {}
        for alias_value, sku in conn.execute(
            """
            SELECT alias_value, sku
            FROM dim_sku_alias
            WHERE alias_type = 'product_name_cn'
              AND is_unique_mapping = 1
            """
        ).fetchall():
            alias_map.setdefault(alias_value, []).append(sku)

        manual_alias_map = {
            alias_value: sku
            for alias_value, sku in conn.execute(
                """
                SELECT alias_value, sku
                FROM manual_sku_alias
                WHERE alias_type = 'product_name_cn'
                  AND is_active = 1
                """
            ).fetchall()
        }

        insert_rows = []
        pending_rows = []
        for row in raw_rows:
            seq, operator, platform, order_date, amazon_order_id, customer, product_name, currency, sale_amount, review_cost = row[:10]
            order_id_text = normalize_text(amazon_order_id)
            if not order_id_text:
                continue

            product_name_text = normalize_text(product_name)
            alias_value = product_name_text.lower()
            mapped_skus = alias_map.get(alias_value, [])
            manual_sku = manual_alias_map.get(alias_value)
            source_skus = source_order_skus(conn, order_id_text)
            fuzzy_skus = fuzzy_product_name_skus(conn, product_name_text)
            sku = None

            if manual_sku:
                sku = manual_sku
            elif len(mapped_skus) == 1:
                sku = mapped_skus[0]
            elif len(source_skus) == 1:
                sku = source_skus[0]
            elif len(fuzzy_skus) == 1:
                sku = fuzzy_skus[0]
            elif len(mapped_skus) > 1:
                pending_rows.append(
                    (
                        'fact_review_orders',
                        str(source_path),
                        None,
                        product_name_text,
                        'product_name_cn',
                        'pending',
                        f'Ambiguous review order mapping for order {order_id_text}',
                        utc_now_iso(),
                        None,
                    )
                )
                continue
            else:
                pending_rows.append(
                    (
                        'fact_review_orders',
                        str(source_path),
                        None,
                        product_name_text,
                        'product_name_cn',
                        'pending',
                        f'No SKU mapping found for review order {order_id_text}',
                        utc_now_iso(),
                        None,
                    )
                )
                continue

            normalized_row = {
                'seq': seq,
                'operator': operator,
                'platform': platform,
                'order_date': parse_order_date(order_date),
                'amazon_order_id': order_id_text,
                'customer': customer,
                'product_name': product_name_text,
                'currency': normalize_text(currency) or None,
                'sale_amount': parse_float(sale_amount),
                'review_cost': parse_float(review_cost),
                'sku': sku,
            }
            row_hash = sha256_text(stable_json(normalized_row))
            insert_rows.append(
                (
                    str(source_path),
                    row_hash,
                    normalized_row['amazon_order_id'],
                    normalized_row['order_date'],
                    normalized_row['product_name'],
                    normalized_row['sku'],
                    normalize_text(platform) or None,
                    normalized_row['currency'],
                    normalized_row['sale_amount'],
                    normalized_row['review_cost'],
                    utc_now_iso(),
                )
            )

        conn.executemany(
            """
            INSERT INTO fact_review_orders (
                source_file,
                source_row_hash,
                amazon_order_id,
                order_date,
                product_name,
                sku,
                platform,
                currency,
                sale_amount,
                review_cost,
                created_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            insert_rows,
        )

        if pending_rows:
            conn.executemany(
                """
                INSERT INTO pending_mapping_queue (
                    source_table,
                    source_file,
                    source_row_hash,
                    ambiguous_value,
                    mapping_type,
                    status,
                    notes,
                    created_at,
                    resolved_at
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                pending_rows,
            )

        conn.commit()
        note = f'Loaded {len(insert_rows)} review orders; pending mappings={len(pending_rows)}'
        finish_etl_run(conn, run_id, 'success', note)
        conn.commit()
        print_banner(note)
        return 0
    except Exception as exc:  # pragma: no cover
        conn.rollback()
        finish_etl_run(conn, run_id, 'failed', str(exc))
        conn.commit()
        raise
    finally:
        conn.close()


if __name__ == '__main__':
    raise SystemExit(main())
