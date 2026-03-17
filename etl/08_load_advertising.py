from __future__ import annotations

import argparse
from pathlib import Path

from openpyxl import load_workbook

from common import connect, finish_etl_run, get_config, print_banner, record_file_import, register_etl_run, utc_now_iso


FILE_PATTERN = '4_Ad Spend_Amazon{yyyymm}.xlsx'
CANDIDATE_SHEETS = ['Sponsored_Products_Advertised_p', '商品推广_推广的商品_报告']
FALLBACK_SUMMARY_SHEETS = ['Sheet2']


def normalize_month(value: str) -> tuple[str, str]:
    text = value.strip()
    if len(text) == 7 and '-' in text:
        return text, text.replace('-', '')
    if len(text) == 6 and text.isdigit():
        return f"{text[:4]}-{text[4:6]}", text
    raise ValueError('target_month must be YYYY-MM or YYYYMM')


def normalize_text(value: object) -> str:
    return str(value).strip() if value is not None else ''


def to_float(value: object) -> float:
    if value is None or value == '':
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    return float(str(value).strip().replace(',', ''))


def get_alias_map(conn):
    alias_map = {}
    for alias_value, sku in conn.execute("select alias_value, sku from dim_sku_alias where alias_type='product_name_cn'"):
        alias_map.setdefault(alias_value, []).append(sku)
    return alias_map


def resolve_sku(conn, alias_map, product_name: str, advertised_sku: str) -> str | None:
    if advertised_sku:
        return advertised_sku
    matches = alias_map.get(product_name.lower(), [])
    if len(matches) == 1:
        return matches[0]
    return None


def load_detailed_sheet(ws, conn, alias_map):
    header = [normalize_text(c) for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    idx = {name: pos for pos, name in enumerate(header)}
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        product_name = normalize_text(row[idx.get('产品', 0)] if '产品' in idx else None)
        advertised_sku = normalize_text(row[idx.get('Advertised SKU')] if 'Advertised SKU' in idx else row[idx.get('广告SKU')] if '广告SKU' in idx else None)
        spend = to_float(row[idx.get('Spend')] if 'Spend' in idx else row[idx.get('花费')] if '花费' in idx else 0)
        impressions = to_float(row[idx.get('Impressions')] if 'Impressions' in idx else row[idx.get('展示量')] if '展示量' in idx else 0)
        clicks = to_float(row[idx.get('Clicks')] if 'Clicks' in idx else row[idx.get('点击量')] if '点击量' in idx else 0)
        sales_7d = to_float(row[idx.get('7 Day Total Sales ')] if '7 Day Total Sales ' in idx else row[idx.get('7天总销售额')] if '7天总销售额' in idx else 0)
        sku = resolve_sku(conn, alias_map, product_name, advertised_sku)
        if not sku:
            continue
        rows.append((sku, spend, impressions, clicks, sales_7d))
    return rows


def load_summary_sheet(ws, alias_map):
    rows = []
    for row in ws.iter_rows(min_row=1, values_only=True):
        product_name = normalize_text(row[0] if len(row) > 0 else None)
        if product_name in ('', '产品'):
            continue
        spend = to_float(row[1] if len(row) > 1 else 0)
        matches = alias_map.get(product_name.lower(), [])
        if len(matches) != 1:
            continue
        rows.append((matches[0], spend, None, None, None))
    return rows


def main() -> int:
    parser = argparse.ArgumentParser(description='Load monthly advertising spend by SKU.')
    parser.add_argument('target_month')
    args = parser.parse_args()
    month_text, month_compact = normalize_month(args.target_month)
    config = get_config()
    source_path = config.base_dir / FILE_PATTERN.format(yyyymm=month_compact)
    if not source_path.exists():
        raise FileNotFoundError(source_path)

    print_banner(f'Loading advertising from {source_path.name}')
    conn = connect(config.db_path)
    run_id = register_etl_run(conn, '08_load_advertising.py', 'load_advertising', target_month=month_text, status='started')
    try:
        wb = load_workbook(source_path, data_only=True, read_only=True)
        record_file_import(conn, run_id, source_path, 'advertising', import_status='loaded', row_count=0)
        alias_map = get_alias_map(conn)
        raw = []
        used_sheet = None
        for sheet in CANDIDATE_SHEETS:
            if sheet in wb.sheetnames:
                used_sheet = sheet
                raw = load_detailed_sheet(wb[sheet], conn, alias_map)
                break
        if not raw:
            for sheet in FALLBACK_SUMMARY_SHEETS:
                if sheet in wb.sheetnames:
                    used_sheet = sheet
                    raw = load_summary_sheet(wb[sheet], alias_map)
                    if raw:
                        break

        agg = {}
        for sku, spend, impressions, clicks, sales_7d in raw:
            current = agg.setdefault(sku, [0.0, 0.0, 0.0, 0.0])
            current[0] += spend or 0.0
            current[1] += impressions or 0.0
            current[2] += clicks or 0.0
            current[3] += sales_7d or 0.0

        conn.execute("DELETE FROM fact_advertising_monthly_sku WHERE source_file = ? AND period_month = ?", (str(source_path), month_text))
        rows = [
            (str(source_path), month_text, sku, vals[0], vals[1], vals[2], vals[3], f'sheet={used_sheet}', utc_now_iso())
            for sku, vals in agg.items()
        ]
        conn.executemany(
            """
            INSERT INTO fact_advertising_monthly_sku (
                source_file, period_month, sku, spend, impressions, clicks, sales_7d, source_note, created_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            rows,
        )
        conn.commit()
        note = f'Loaded {len(rows)} advertising SKU rows for {month_text} from {used_sheet}'
        finish_etl_run(conn, run_id, 'success', note)
        conn.commit()
        print_banner(note)
        return 0
    except Exception as exc:
        conn.rollback()
        finish_etl_run(conn, run_id, 'failed', str(exc))
        conn.commit()
        raise
    finally:
        conn.close()


if __name__ == '__main__':
    raise SystemExit(main())
